from flask import Flask, request, jsonify, session, render_template, redirect, url_for, Response
from werkzeug.security import check_password_hash, generate_password_hash
from database import get_db, init_db
from functools import wraps
import os, io

app = Flask(__name__)
app.secret_key = os.urandom(24)

# ─── Auth Decorators ──────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        if session.get('role') != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated

# ─── Page Routes ──────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login_page'))
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('user_logbook'))

@app.route('/login')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if session.get('role') != 'admin':
        return redirect(url_for('user_logbook'))
    return render_template('admin/dashboard.html', username=session.get('username'))

@app.route('/admin/machines')
@login_required
def admin_machines():
    if session.get('role') != 'admin':
        return redirect(url_for('user_logbook'))
    return render_template('admin/machines.html', username=session.get('username'))

@app.route('/admin/logbook')
@login_required
def admin_logbook():
    if session.get('role') != 'admin':
        return redirect(url_for('user_logbook'))
    return render_template('admin/logbook.html', username=session.get('username'))

@app.route('/admin/analysis')
@login_required
def admin_analysis():
    if session.get('role') != 'admin':
        return redirect(url_for('user_logbook'))
    return render_template('admin/analysis.html', username=session.get('username'))

@app.route('/admin/settings')
@login_required
def admin_settings():
    if session.get('role') != 'admin':
        return redirect(url_for('user_logbook'))
    return render_template('admin/settings.html', username=session.get('username'))

@app.route('/user/logbook')
@login_required
def user_logbook():
    return render_template('user/logbook.html',
                           username=session.get('username'),
                           role=session.get('role'))

# ─── Auth API ─────────────────────────────────────────────────────────────────

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
    db.close()
    if not user or not check_password_hash(user['password_hash'], password):
        return jsonify({'error': 'Invalid username or password'}), 401
    session['user_id'] = user['id']
    session['username'] = user['username']
    session['role'] = user['role']
    redirect_url = '/admin/dashboard' if user['role'] == 'admin' else '/user/logbook'
    return jsonify({'success': True, 'role': user['role'], 'redirect': redirect_url})

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True})

@app.route('/api/auth/me')
def api_me():
    if 'user_id' not in session:
        return jsonify({'authenticated': False}), 401
    return jsonify({'authenticated': True, 'username': session['username'], 'role': session['role']})

# ─── Users API ────────────────────────────────────────────────────────────────

@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    db = get_db()
    users = db.execute("SELECT id, username, role, created_at FROM users ORDER BY created_at DESC").fetchall()
    db.close()
    return jsonify([dict(u) for u in users])

@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    role = data.get('role', 'user')
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    if role not in ('admin', 'user'):
        return jsonify({'error': 'Invalid role'}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                   (username, generate_password_hash(password), role))
        db.commit()
    except Exception:
        db.close()
        return jsonify({'error': 'Username already exists'}), 409
    db.close()
    return jsonify({'success': True})

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
def delete_user(user_id):
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    if user_id == session['user_id']:
        return jsonify({'error': 'Cannot delete yourself'}), 400
    db = get_db()
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()
    db.close()
    return jsonify({'success': True})

# ─── Machines API ─────────────────────────────────────────────────────────────
# Fields: equipment_no, name, asset_code, location, area, module, capacity, unit, make, installation_date

@app.route('/api/machines', methods=['GET'])
@login_required
def get_machines():
    db = get_db()
    rows = db.execute("SELECT * FROM machines ORDER BY equipment_no").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/machines', methods=['POST'])
@login_required
def create_machine():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    eq_no = data.get('equipment_no', '').strip()
    name  = data.get('name', '').strip()
    if not eq_no or not name:
        return jsonify({'error': 'Equipment No and Equipment Name are required'}), 400
    db = get_db()
    try:
        db.execute(
            'INSERT INTO machines (equipment_no, name, asset_code, location, area, module, capacity, unit, make, installation_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (eq_no, name,
             data.get('asset_code', ''), data.get('location', ''),
             data.get('area', ''),       data.get('module', ''),
             data.get('capacity', ''),   data.get('unit', ''),
             data.get('make', ''),       data.get('installation_date', ''))
        )
        db.commit()
    except Exception:
        db.close()
        return jsonify({'error': 'Equipment No already exists'}), 409
    db.close()
    return jsonify({'success': True})

@app.route('/api/machines/<int:machine_id>', methods=['PUT'])
@login_required
def update_machine(machine_id):
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    db = get_db()
    db.execute(
        'UPDATE machines SET equipment_no=?, name=?, asset_code=?, location=?, area=?, module=?, capacity=?, unit=?, make=?, installation_date=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
        (data.get('equipment_no'), data.get('name'),
         data.get('asset_code', ''), data.get('location', ''),
         data.get('area', ''),       data.get('module', ''),
         data.get('capacity', ''),   data.get('unit', ''),
         data.get('make', ''),       data.get('installation_date', ''),
         machine_id)
    )
    db.commit()
    db.close()
    return jsonify({'success': True})

@app.route('/api/machines/<int:machine_id>', methods=['DELETE'])
@login_required
def delete_machine(machine_id):
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    db = get_db()
    db.execute("DELETE FROM machines WHERE id = ?", (machine_id,))
    db.commit()
    db.close()
    return jsonify({'success': True})

# ─── Dropdown Options API ─────────────────────────────────────────────────────

@app.route('/api/dropdown-options', methods=['GET'])
@login_required
def get_dropdown_options():
    field = request.args.get('field', '')
    db = get_db()
    if field:
        rows = db.execute("SELECT * FROM dropdown_options WHERE field_name=? ORDER BY display_order, value", (field,)).fetchall()
    else:
        rows = db.execute("SELECT * FROM dropdown_options ORDER BY field_name, display_order, value").fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/dropdown-options', methods=['POST'])
@login_required
def create_dropdown_option():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    field = data.get('field_name', '').strip()
    value = data.get('value', '').strip()
    if not field or not value:
        return jsonify({'error': 'field_name and value required'}), 400
    db = get_db()
    try:
        db.execute("INSERT INTO dropdown_options (field_name, value) VALUES (?, ?)", (field, value))
        db.commit()
    except Exception:
        db.close()
        return jsonify({'error': 'Option already exists'}), 409
    db.close()
    return jsonify({'success': True})

@app.route('/api/dropdown-options/<int:option_id>', methods=['DELETE'])
@login_required
def delete_dropdown_option(option_id):
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    db = get_db()
    db.execute("DELETE FROM dropdown_options WHERE id = ?", (option_id,))
    db.commit()
    db.close()
    return jsonify({'success': True})

# ─── Log Entries API ──────────────────────────────────────────────────────────

@app.route('/api/log-entries', methods=['GET'])
@login_required
def get_log_entries():
    db = get_db()
    query = 'SELECT le.*, m.equipment_no, m.name as machine_name, m.location as machine_location, u.username as submitted_by_name FROM log_entries le JOIN machines m ON le.machine_id = m.id JOIN users u ON le.submitted_by = u.id WHERE 1=1'
    params = []
    for arg, col in [('review_status','le.review_status'),('status','le.status'),('category','le.category'),('machine_id','le.machine_id')]:
        v = request.args.get(arg)
        if v: query += f' AND {col} = ?'; params.append(v)
    date_from = request.args.get('date_from')
    if date_from: query += ' AND le.entry_date >= ?'; params.append(date_from)
    date_to = request.args.get('date_to')
    if date_to: query += ' AND le.entry_date <= ?'; params.append(date_to)
    if session.get('role') != 'admin':
        query += ' AND le.submitted_by = ?'; params.append(session['user_id'])
    query += ' ORDER BY le.created_at DESC'
    rows = db.execute(query, params).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/log-entries', methods=['POST'])
@login_required
def create_log_entry():
    data = request.get_json()
    required = ['machine_id', 'entry_date', 'start_time', 'end_time', 'trouble_description']
    for field in required:
        if not data.get(field):
            return jsonify({'error': f'{field} is required'}), 400
    try:
        from datetime import datetime
        end_date = data.get('end_date') or data['entry_date']
        start_dt = datetime.strptime(f"{data['entry_date']} {data['start_time']}", '%Y-%m-%d %H:%M')
        end_dt   = datetime.strptime(f"{end_date} {data['end_time']}", '%Y-%m-%d %H:%M')
        diff = (end_dt - start_dt).total_seconds() / 60
        if diff < 0: diff = 0
        down_time = int(diff)
    except Exception:
        end_date  = data.get('end_date') or data['entry_date']
        down_time = 0
    db = get_db()
    db.execute(
        'INSERT INTO log_entries (machine_id, area, module, entry_date, end_date, shift, start_time, end_time, down_time_minutes, error_message, trouble_description, maintenance_action, quality_confirmation, category, occurrence, status, engineer1, engineer2, engineer3, submitted_by, review_status) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, \'pending\')',
        (data['machine_id'], data.get('area',''), data.get('module',''), data['entry_date'],
         end_date, data.get('shift',''), data['start_time'], data['end_time'], down_time,
         data.get('error_message',''), data['trouble_description'], data.get('maintenance_action',''),
         data.get('quality_confirmation',''), data.get('category',''), data.get('occurrence',''),
         data.get('status','Open'), data.get('engineer1',''), data.get('engineer2',''),
         data.get('engineer3',''), session['user_id'])
    )
    db.commit()
    db.close()
    return jsonify({'success': True})

@app.route('/api/log-entries/<int:entry_id>', methods=['PUT'])
@login_required
def update_log_entry(entry_id):
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    data = request.get_json()
    try:
        from datetime import datetime
        end_date2 = data.get('end_date') or data['entry_date']
        start_dt = datetime.strptime(f"{data['entry_date']} {data['start_time']}", '%Y-%m-%d %H:%M')
        end_dt   = datetime.strptime(f"{end_date2} {data['end_time']}", '%Y-%m-%d %H:%M')
        diff = (end_dt - start_dt).total_seconds() / 60
        if diff < 0: diff = 0
        down_time = int(diff)
    except Exception:
        end_date2 = data.get('end_date') or data.get('entry_date','')
        down_time = data.get('down_time_minutes', 0)
    review_status = 'confirmed' if data.get('confirm') else 'pending'
    db = get_db()
    db.execute(
        'UPDATE log_entries SET machine_id=?, area=?, module=?, entry_date=?, end_date=?, shift=?, start_time=?, end_time=?, down_time_minutes=?, error_message=?, trouble_description=?, maintenance_action=?, quality_confirmation=?, category=?, occurrence=?, status=?, engineer1=?, engineer2=?, engineer3=?, checked_by=?, review_status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?',
        (data['machine_id'], data.get('area',''), data.get('module',''), data['entry_date'],
         end_date2, data.get('shift',''), data['start_time'], data['end_time'], down_time,
         data.get('error_message',''), data['trouble_description'], data.get('maintenance_action',''),
         data.get('quality_confirmation',''), data.get('category',''), data.get('occurrence',''),
         data.get('status','Open'), data.get('engineer1',''), data.get('engineer2',''),
         data.get('engineer3',''), data.get('checked_by',''), review_status, entry_id)
    )
    db.commit()
    db.close()
    return jsonify({'success': True})

@app.route('/api/log-entries/<int:entry_id>', methods=['DELETE'])
@login_required
def delete_log_entry(entry_id):
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    db = get_db()
    db.execute("DELETE FROM log_entries WHERE id = ?", (entry_id,))
    db.commit()
    db.close()
    return jsonify({'success': True})

# ─── Export Log Entries as Excel ──────────────────────────────────────────────

@app.route('/api/log-entries/export', methods=['GET'])
@login_required
def export_log_entries():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    query = 'SELECT le.*, m.equipment_no, m.name as machine_name, m.location as machine_location, u.username as submitted_by_name FROM log_entries le JOIN machines m ON le.machine_id = m.id JOIN users u ON le.submitted_by = u.id WHERE 1=1'
    params = []
    for arg, col in [('status','le.status'),('category','le.category'),('machine_id','le.machine_id'),('review_status','le.review_status')]:
        v = request.args.get(arg)
        if v: query += f' AND {col} = ?'; params.append(v)
    date_from = request.args.get('date_from')
    if date_from: query += ' AND le.entry_date >= ?'; params.append(date_from)
    date_to = request.args.get('date_to')
    if date_to: query += ' AND le.entry_date <= ?'; params.append(date_to)
    query += ' ORDER BY le.entry_date DESC, le.created_at DESC'
    rows = db.execute(query, params).fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Maintenance Log'

    # Header styling
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='2E9B3E')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='B0C4B8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'Start Date', 'Equipment No', 'Machine Name', 'Location', 'Area', 'Module',
               'Shift', 'Start Time', 'End Date', 'End Time', 'Down Time (hrs)',
               'Error Message', 'Trouble Description', 'Maintenance Action', 'Quality Confirm',
               'Category', 'Occurrence', 'Status', 'Engineer 1', 'Engineer 2', 'Engineer 3',
               'Submitted By', 'Checked By', 'Review Status']
    col_widths = [4, 12, 14, 20, 14, 12, 14, 10, 10, 12, 10, 14, 20, 30, 28, 14, 14, 18, 14, 14, 14, 14, 14, 14, 14]

    ws.row_dimensions[1].height = 28
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Alternate row fill
    fill_even = PatternFill('solid', fgColor='EAF7EC')
    data_align = Alignment(vertical='center', wrap_text=False)

    for ri, row in enumerate(rows, 2):
        r = dict(row)
        fill = fill_even if ri % 2 == 0 else None
        dt_mins = r.get('down_time_minutes', 0) or 0
        dt_h = dt_mins // 60; dt_m = dt_mins % 60
        dt_str = (f'{dt_h}h {dt_m}m' if dt_m else f'{dt_h}h') if dt_h else f'{dt_m}m'
        end_dt_val = r.get('end_date') or r.get('entry_date','')
        values = [ri-1, r.get('entry_date',''), r.get('equipment_no',''), r.get('machine_name',''),
                  r.get('machine_location',''), r.get('area',''), r.get('module',''),
                  (r.get('shift','') + ' Shift') if r.get('shift') else '',
                  r.get('start_time',''), end_dt_val, r.get('end_time',''), dt_str,
                  r.get('error_message',''), r.get('trouble_description',''), r.get('maintenance_action',''),
                  r.get('quality_confirmation',''), r.get('category',''), r.get('occurrence',''),
                  r.get('status',''), r.get('engineer1',''), r.get('engineer2',''), r.get('engineer3',''),
                  r.get('submitted_by_name',''), r.get('checked_by',''), r.get('review_status','')]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = data_align; cell.border = border
            if fill: cell.fill = fill

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = 'maintenance_log.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

# ─── Analysis API ─────────────────────────────────────────────────────────────

@app.route('/api/analysis', methods=['GET'])
@login_required
def get_analysis():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    db = get_db()
    params = []
    where = "WHERE 1=1"
    for arg, col in [('date_from','le.entry_date >='),('date_to','le.entry_date <='),('machine_id','le.machine_id ='),('category','le.category ='),('status','le.status =')]:
        v = request.args.get(arg.split('_')[0] if arg == 'date_from' else (arg.split('_')[0] if arg == 'date_to' else arg))
        v = request.args.get(arg)
        if v: where += f' AND {col} ?'; params.append(v)
    pareto_machine = db.execute(f'SELECT m.equipment_no, m.name as machine_name, SUM(le.down_time_minutes) as total_downtime, COUNT(le.id) as incident_count FROM log_entries le JOIN machines m ON le.machine_id = m.id {where} GROUP BY le.machine_id ORDER BY total_downtime DESC LIMIT 10', params).fetchall()
    pareto_category = db.execute(f'SELECT le.category, SUM(le.down_time_minutes) as total_downtime, COUNT(le.id) as incident_count FROM log_entries le JOIN machines m ON le.machine_id = m.id {where} GROUP BY le.category ORDER BY total_downtime DESC', params).fetchall()
    mttr_row = db.execute(f'SELECT AVG(le.down_time_minutes) as avg_downtime, COUNT(le.id) as total_incidents, SUM(le.down_time_minutes) as total_downtime FROM log_entries le JOIN machines m ON le.machine_id = m.id {where}', params).fetchone()
    monthly_trend = db.execute(f'SELECT strftime(\'%Y-%m\', le.entry_date) as month, SUM(le.down_time_minutes) as total_downtime, COUNT(le.id) as incident_count FROM log_entries le JOIN machines m ON le.machine_id = m.id {where} GROUP BY month ORDER BY month', params).fetchall()
    status_breakdown = db.execute(f'SELECT le.status, COUNT(le.id) as count FROM log_entries le JOIN machines m ON le.machine_id = m.id {where} GROUP BY le.status', params).fetchall()
    db.close()
    return jsonify({'pareto_machine':[dict(r) for r in pareto_machine],'pareto_category':[dict(r) for r in pareto_category],'mttr':dict(mttr_row) if mttr_row else {},'monthly_trend':[dict(r) for r in monthly_trend],'status_breakdown':[dict(r) for r in status_breakdown]})

# ─── Dashboard Stats ──────────────────────────────────────────────────────────

@app.route('/api/dashboard-stats')
@login_required
def dashboard_stats():
    if session.get('role') != 'admin': return jsonify({'error':'Admin only'}), 403
    db = get_db()
    tm = db.execute('SELECT COUNT(*) FROM machines').fetchone()[0]
    te = db.execute('SELECT COUNT(*) FROM log_entries').fetchone()[0]
    pe = db.execute("SELECT COUNT(*) FROM log_entries WHERE review_status='pending'").fetchone()[0]
    oi = db.execute("SELECT COUNT(*) FROM log_entries WHERE status='Open'").fetchone()[0]
    recent = db.execute('SELECT le.id,le.entry_date,le.trouble_description,le.review_status,le.down_time_minutes,m.equipment_no,m.name as machine_name,u.username as submitted_by_name FROM log_entries le JOIN machines m ON le.machine_id=m.id JOIN users u ON le.submitted_by=u.id ORDER BY le.created_at DESC LIMIT 5').fetchall()
    db.close()
    return jsonify({'total_machines':tm,'total_entries':te,'pending_entries':pe,'open_issues':oi,'recent_entries':[dict(r) for r in recent]})

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
